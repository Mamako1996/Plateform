import time
import cv2
import pytesseract
import openpyxl
import re

# new Excel file making
from openpyxl import load_workbook

wb = load_workbook('Testing.xlsx')
#wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Testing Data'
#sheet.append(['Number', 'UV Transmission', 'IR Transmission', 'VL Transmission'])

# Tesseract-OCR path
pytesseract.pytesseract.tesseract_cmd = pytesseract.pytesseract.tesseract_cmd = r"C:\Program " \
                                                                                r"Files\Tesseract-OCR\tesseract.exe "
cap = cv2.VideoCapture(1, cv2.CAP_DSHOW)
signal = cap.isOpened()
index = 1

# open Tesseract
temp = ''
ret, frame = cap.read()
cv2.imshow("Current", frame)
# Image open
img = cv2.imread('Test11.jpg')
img = img[0:750, 200:460]

# Resizing for scan
img = cv2.resize(img, None, fx=1.1, fy=1.1)
# threshold method for dividing the clearer color block
tn = 0
ti = 180
start = time.perf_counter()
while True:
    ret, r1 = cv2.threshold(img, ti + tn, 255, cv2.THRESH_BINARY)
    rz = cv2.cvtColor(r1, cv2.COLOR_BGR2GRAY)
    ret, rr = cv2.threshold(rz, 0, 255, cv2.THRESH_BINARY)
    #print(tn)
    # extract the text from image
    text = pytesseract.image_to_string(rr)
    text = text.replace(" ", "")
    text = text.replace(",", ".")

    testchar = 0

    # make a cell to hold the data
    a = re.findall(r"\d+\.?\d*", text)
    if len(a) == 3:
        for n in range(3):
            if a[n].find("100") != -1:
                a[n] = '100'
            if len(a[n]) > 4:
                a[n] = a[n][0:2]
            a[n] = str(round(float(a[n]), 1))
        for line in text.splitlines():
            b = re.findall(r'\d+', line)
            if len(b) != 0:
                if float(b[0]) > 100:
                    testchar = 1
                else:
                    if line.find("%") != -1:
                        for c in line:
                            if c.isalpha():
                                testchar = 1
                        if (line.find(".") == -1) & (line.find("100") == -1):  # & (line.find("100") == -1)
                            testchar = 1
                    else:
                        if a[n].find("100") == -1:
                            testchar = 1
            else:
                if line.find("%") != -1:
                    testchar = 1
        if testchar == 0:
            # print it out for checking
            print(text + "\n")
            print("threshold value is " + str(tn + ti))
            break
        else:
            tn += 1
            testchar = 0
    else:
        tn += 1
    if (ti + tn) >= 255:
        break
print(a)
sheet.append(a)
wb.save('Testing.xlsx')
end = time.perf_counter()
print("time consuming : {:.2f}s".format(end - start))
#
# cv2.imshow('Result0', img)
# cv2.imshow('Result1', r1)
# cv2.imshow('Result2', rz)
# cv2.imshow('Result3', rr)
cv2.waitKey(0)